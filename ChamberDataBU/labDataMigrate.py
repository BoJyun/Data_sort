# -*- coding: utf-8 -*-
"""
Created on Tue Oct  6 18:02:07 2020

@author: Johnny_Huang
"""

from openpyxl import load_workbook
import win32com.client

def datamigrate(name,namelab):
    wb = load_workbook(r'D:\\labData\\'+name)
    sheet = wb.active
    a=0
    for row1 in range(4,sheet.max_row):
        if (sheet.cell(row=row1, column=1).value==None):
            a=row1
            break

    wb2= load_workbook(r'D:\\labData\\'+namelab)
    #print(wb2.sheetnames)
    sheet2 = wb2['SG64']
    for i in range(4,a):
        for j in range(1, 8):
            sheet2.cell(row=i, column=j).value=sheet.cell(row=i, column=j).value

    wb2.save(r'D:\\labData\\'+namelab)
    wb2.close()
    wb.close()

    excel=win32com.client.Dispatch('Excel.Application')
    excel.Visible=False
    myBook=excel.Workbooks.open(r'D:\\labData\\'+namelab)
    myBook.Save
    myBook.Close(True)
    excel.Application.Quit()

def changeWeekNum(name):
    # wb = load_workbook(r'D:\\labData\\'+name)
    # sheet = wb.active
    # sheet.cell(row=2, column=7).value=sheet.cell(row=2, column=7).value+1
    # print(sheet.cell(row=2, column=7).value)
    # wb.save(r'D:\\labData\\'+name)
    # wb.close()

    excel=win32com.client.Dispatch('Excel.Application')
    excel.Visible=False
    myBook=excel.Workbooks.open(r'D:\\labData\\'+name)
    xlSheet = myBook.Worksheets('Day&Night Week Chart Chamber')
    xlSheet.Cells(6,17).Value = int(xlSheet.Cells(6,17).Value)+1
    myBook.Save
    myBook.Close(True)
    excel.Application.Quit()
    # xlSheet.Cells(2,1).Value = 123

    # excel=win32com.client.Dispatch('Excel.Application')
    # excel.Visible=True
    # myBook=excel.Workbooks.open(r'D:\\labData\\Graph_Correct_new2.xlsx')
    # myBook.Save
    # myBook.Close(True)
    # excel.Application.Quit()

if __name__=='__main__':
    datamigrate('ETS8500 Utilization Rate.xlsx','ETS8500lab Utilization Rate.xlsx')
    datamigrate('Aten Utilization Rate.xlsx','Atenlab Utilization Rate.xlsx')
    datamigrate('CATR Utilization Rate.xlsx','CATRlab Utilization Rate.xlsx')
    datamigrate('SG24 Utilization Rate.xlsx','SG24lab Utilization Rate.xlsx')
    datamigrate('SG64 Utilization Rate.xlsx','SG64lab Utilization Rate.xlsx')
    datamigrate('TP Utilization Rate.xlsx','TPlab Utilization Rate.xlsx')
    datamigrate('RCARD Utilization Rate.xlsx','R-cardlab Utilization Rate.xlsx')
    changeWeekNum('Graph_Correct_new2.xlsx')

